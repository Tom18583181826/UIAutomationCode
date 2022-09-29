from openpyxl import load_workbook
# openpyxl是Python的一个第三方库，用于处理excel表格
from common.read_ini import ReadIni
from common.read_json import ReadJson
from data.excel_column import ExcelColumn


class ReadExcel:
    def __init__(self):
        self.read_ini = ReadIni()
        self.excel_path = self.read_ini.get_excel_file_path()
        # 获取excel文件路径
        self.params_file_path = self.read_ini.get_json_file_path()
        # 获取参数文件路径
        self.expect_file_path = self.read_ini.get_json_file_path()
        # 获取预期结果文件路径
        self.wb = load_workbook(self.excel_path)
        # 获取excel表格
        self.wb_sheet = self.wb[self.read_ini.get_sheet_name()]
        # 获取sheet页

    # 获取指定单元格的值，传入的参数值为excel表的行和列
    def get_cell_value(self, column, row):
        return self.wb_sheet[column + str(row)].value

    # 获取用例编号，传入的参数值为excel表的行，列指定写死
    def get_case_id(self, row):
        return self.get_cell_value(ExcelColumn.CASE_ID, row)

    # 获取用例标题
    def get_case_name(self, row):
        return self.get_cell_value(ExcelColumn.CASE_NAME, row)

    # 获取参数的key
    def get_case_params_key(self, row):
        return self.get_cell_value(ExcelColumn.CASE_PARAMS_KEY, row)

    # 获取参数的value
    def get_case_params_value(self, row):
        if self.get_case_params_key(row):
            return ReadJson().read_json_data(self.params_file_path)[self.get_case_params_key(row)]

    # 获取预期结果的key
    def get_case_expect_key(self, row):
        return self.get_cell_value(ExcelColumn.CASE_EXPECT_KEY, row)

    # 获取预期结果的value
    def get_case_expect_value(self, row):
        if self.get_case_expect_key(row):
            return ReadJson().read_json_data(self.expect_file_path)[self.get_case_expect_key(row)]

    # 获取excel表的行数
    def get_row_count(self):
        return self.wb_sheet.max_row


if __name__ == '__main__':
    read = ReadExcel()
    data_list = []
    for row in range(2, read.get_row_count() + 1):
        id = read.get_case_id(row)
        name = read.get_case_name(row)
        params = read.get_case_params_value(row)
        expect = read.get_case_expect_value(row)
        data_list.append((id, name, params, expect))
    print(data_list)
